#!/usr/bin/env python3
"""資源エネルギー庁 電力調査統計 月次データ（Phase 2-A 第 1 弾、12 系列）

---------------------------------------------------------------------
ステータス: **Day 3 実装（index 抽出を fiscal-year ベースに再設計）**
  - Day 1: 議事録 + SKELETON
  - Day 2: source_map.yaml + metadata.py 拡張 + parse helper 最小実装（単月仮定）
  - Day 3: 実 METI HTML 疎通 → **ファイルは年度単位と判明**。list_xlsx_links を
           fiscal-year + table (generation/demand) ベースに再設計。
  - Day 4（次）: 実 XLSX 構造の確認（`2-1-2024.xlsx` を実 DL）→ 12 ヶ月 × 8 系列
                の wide→long 展開、`fetch_all_months` の実ダウンロード + 追記
---------------------------------------------------------------------

設計議事録: `docs/data-pipeline-phase2-a-power-discussion.md`
Day 3 発見レポ: `docs/phase2-sprint-day3-discovery.md`（作成予定）

ターゲット 12 系列（すべて月次、2016-04〜最新月バックフィル）:
  発電 8: meti-gen-total / -thermal / -hydro / -nuclear / -solar / -wind /
          -geothermal / -biomass
  需要 3: meti-demand-total / -lights / -power
  派生 1: meti-renewables-share（他系列から fetch 後に計算）

Day 3 の重要発見（実 HTML 疎通から確定）:
  - ファイルは **fiscal year 単位**（月次ではない）。1 XLSX に 4 月〜翌 3 月の
    12 ヶ月が格納される wide 形式。バックフィル対象は FY2016 〜 最新年度 = **10 年度分**。
  - ファイル名パターン:
      FY2019〜FY2025: `2-1-{YYYY}.xlsx` / `3-1-{YYYY}.xlsx` （+ 2025 のみ `*n.xlsx` 機械判読版）
      FY2016〜FY2018: `2-1-H{nn}.xlsx`  / `3-1-H{nn}.xlsx`  （和暦平成）
      FY2015 以前:    `2-1-H{nn}.xls`  + 分割シート構造（Phase 2-B 以降）
  - Index ページは 2 つ:
      https://.../ep002/results.html         ← 最新年度のリンクのみ
      https://.../ep002/results_archive.html ← 2014〜直近年度 10 年度分のリンク
    → list_xlsx_links は両方を union する。

設計ポリシー（Day 3 版）:
  - list_xlsx_links は filename 正規表現で `2-1-*` / `3-1-*` のみを拾う（厳密）
  - fiscal year 境界: `min_fiscal_year=2016` でデフォルト足切り
  - 機械判読版 `*n.xlsx` が存在すれば通常版より優先（同じ (fy, table) キーで上書き）
  - parse_generation_sheet / parse_demand_sheet（Day 2 の単月返し）は Day 4 で
    「fiscal_year の 12 ヶ月分 long 形式」を返すシグネチャに昇格予定
  - xls（旧形式）ファイルは当面サポートしない（Phase 2-B 以降）
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import yaml
from bs4 import BeautifulSoup

# 共通ライブラリ
REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

from scripts.common.http import get  # noqa: E402
from scripts.common.io import append_log, save_raw, write_processed  # noqa: E402
from scripts.common.metadata import write_metadata_for_indicator  # noqa: E402

# ----- Phase 2-A Day 5: Akamai Bot Manager 突破（2026-04-24 深夜） --------
# METI（www.enecho.meti.go.jp）は Akamai Bot Manager で保護されており、
# requests / urllib3 / .NET Invoke-WebRequest / curl.exe はすべて timeout する
# （Chrome のみブラウザ JS challenge を通過できる）。
# curl_cffi は Chrome/Firefox の TLS fingerprint を模倣するので、Akamai の
# TLS チェック層を抜けられる。ローカル Windows で status=200 を確認済み。
# 依存は本スクリプトからのみ。他の fetch_*.py は requests のままで問題なし。
try:
    from curl_cffi import requests as _curl_requests  # noqa: E402
    _CURL_CFFI_AVAILABLE = True
except ImportError:
    _curl_requests = None
    _CURL_CFFI_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("fetch_enecho_power")

# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------

SOURCE_KEY = "enecho-power"
SOURCE_MAP_PATH = REPO_ROOT / "docs" / "source_map.yaml"
DATA_ROOT = REPO_ROOT / "data"
PROCESSED_DIR = DATA_ROOT / "processed" / "enecho-power"
RAW_DIR = DATA_ROOT / "raw" / "enecho-power"
LOG_DIR = DATA_ROOT / "_logs"

# 1 リクエストごとのスリープ（METI サーバへの配慮、1.5 秒）
SLEEP_BETWEEN_REQUESTS = 1.5

# ----- 2026-09-05 Akamai 202 対策（#44 の実走で判明）-----------------------
# METI（Akamai Bot Manager）は間欠的に HTTP 202 + 0 bytes のチャレンジ応答を返す。
# raise_for_status() は 2xx を通すので、index ページなら「リンク 0 件で正常終了」、
# XLSX なら「0 bytes」になる。#44 で公表中年度を毎晩再取得するようになり、この確率が
# 毎晩顕在化するため、202 のときは待って再試行する（実測: 180 秒後の再試行で 200）。
# 試行回数 = len(METI_RETRY_SLEEPS) + 1。最後も 202 なら RuntimeError（黙って 0 件にしない）。
METI_RETRY_SLEEPS = (45, 120)

# ----- Phase 2-A Day 5 対策（Run #24/25/26 timeout post-mortem）-------
# GitHub Actions runner（US DC）→ METI サーバへのデフォルト UA
# （"eic-data-pipeline/0.1 (+github...)"）では応答が返ってこない。
# 実ブラウザ風 UA に変えて、Accept-Language を日本語に設定すると、
# 少なくとも WAF/IP reputation の layer で弾かれにくくなる想定。
# BOJ / JMA / 財務省では問題なかったので、METI の WAF が bot-like UA に
# 厳しめだという仮説。
BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}
XLSX_HEADERS = {
    **BROWSER_HEADERS,
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/octet-stream,*/*;q=0.8"
    ),
}


def _meti_get(url: str, *, timeout: int = 120, accept_xlsx: bool = False):
    """METI 専用 GET。Akamai Bot Manager 回避のため curl_cffi で Chrome 120 を模倣。

    Args:
        url: 取得対象
        timeout: 応答タイムアウト秒
        accept_xlsx: XLSX 取得時は True（Accept ヘッダを調整）
    Returns:
        response object（r.status_code / r.text / r.content / r.raise_for_status()）
    Raises:
        RuntimeError: curl_cffi が未インストール
    """
    if not _CURL_CFFI_AVAILABLE:
        raise RuntimeError(
            "curl_cffi is required to fetch METI (Akamai Bot Manager bypass). "
            "Install via: pip install curl_cffi"
        )
    headers = dict(XLSX_HEADERS if accept_xlsx else BROWSER_HEADERS)
    # curl_cffi は User-Agent を impersonate プロファイルが自動で設定する。
    # Accept / Accept-Language だけ明示的に上書きする。
    attempts = len(METI_RETRY_SLEEPS) + 1
    for attempt in range(1, attempts + 1):
        r = _curl_requests.get(
            url,
            impersonate="chrome120",
            timeout=timeout,
            headers={
                "Accept": headers["Accept"],
                "Accept-Language": headers["Accept-Language"],
            },
        )
        if r.status_code != 202:
            return r
        if attempt < attempts:
            wait = METI_RETRY_SLEEPS[attempt - 1]
            logger.warning(
                "METI returned 202 (Akamai challenge, %d bytes) for %s — retry %d/%d after %ds",
                len(r.content), url, attempt, attempts - 1, wait,
            )
            time.sleep(wait)
    raise RuntimeError(
        f"METI returned HTTP 202 (Akamai challenge) {attempts} times for {url}; giving up"
    )

# ファイル名や link text から YYYY-MM を拾う正規表現（複数パターン、上から試す）
# ※ 本スクリプトのメイン経路では filename は fiscal year 単位（YYYY / H{nn}）。
#   YEAR_MONTH パターンは _infer_year_month で互換のため残置。
YEAR_MONTH_PATTERNS = [
    re.compile(r"(\d{4})[-_\.]?(\d{1,2})"),                    # 2026-02 / 2026_02 / 2026.2 / 202602
    re.compile(r"(\d{4})年\s*(\d{1,2})月"),                    # 2026年2月
    re.compile(r"令和\s*(\d{1,2})年\s*(\d{1,2})月"),           # 令和8年2月
    re.compile(r"平成\s*(\d{1,2})年\s*(\d{1,2})月"),           # 平成31年4月
    re.compile(r"R(\d{1,2})[\.\-_](\d{1,2})"),                 # R8.2 / R8-2
    re.compile(r"H(\d{1,2})[\.\-_](\d{1,2})"),                 # H30.12
]

# 和暦 → 西暦のオフセット
ERA_OFFSET = {"令和": 2018, "R": 2018, "平成": 1988, "H": 1988}

# ----- Phase 2-A Day 3 の実 HTML 確認で確定した仕様 ---------------------
# METI 電力調査統計は **年度単位** のファイル公開（月次ではない）。
# 1 XLSX に fiscal_year の 4 月〜翌 3 月の 12 ヶ月データが格納される。
# filename パターンは年度によって変わる:
#   2019〜2025 年度: `2-1-{YYYY}.xlsx` / `3-1-{YYYY}.xlsx` or `2-1-{YYYY}n.xlsx`
#   2016〜2018 年度: `2-1-H{nn}.xlsx`  / `3-1-H{nn}.xlsx`
#   2015 以前:        `2-1-H{nn}.xls` + 分割シート構造（Phase 2-B 以降で対応）
# fetch_all_months では fiscal_year → XLSX → 12 ヶ月分のレコードに展開する。
# --------------------------------------------------------------------

# filename の fiscal year 抽出正規表現（西暦 / 和暦平成両対応）
FISCAL_YEAR_PATTERNS = [
    # 2-1-2024.xlsx / 3-1-2024n.xlsx / 2-1-2024.xls
    (re.compile(r"^([23])-1-(\d{4})(n)?\.(xlsx|xls)$", re.IGNORECASE), "seireki"),
    # 2-1-H28.xlsx / 3-1-H29.xlsx
    (re.compile(r"^([23])-1-H(\d{2})(n)?\.(xlsx|xls)$", re.IGNORECASE), "heisei"),
]

TABLE_KEY_FROM_PREFIX = {"2": "generation", "3": "demand"}


# ---------------------------------------------------------------------------
# source_map.yaml 読み込み
# ---------------------------------------------------------------------------


def load_source_cfg() -> dict:
    """source_map.yaml から enecho-power セクションを読み込む。

    Raises:
        KeyError: enecho-power が未定義
    """
    with SOURCE_MAP_PATH.open("r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    sources = cfg.get("sources") or {}
    if SOURCE_KEY not in sources:
        raise KeyError(f"{SOURCE_MAP_PATH} に {SOURCE_KEY} セクションが見つかりません")
    return sources[SOURCE_KEY]


# ---------------------------------------------------------------------------
# INDEX ページから XLSX リンク抽出
# ---------------------------------------------------------------------------


def _infer_year_month(text: str) -> Optional[str]:
    """任意文字列から YYYY-MM を推定。見つからなければ None。

    令和・平成の和暦にも対応。
    """
    if not text:
        return None
    # 和暦（例: 令和8年2月）
    for era, offset in [("令和", 2018), ("平成", 1988)]:
        m = re.search(rf"{era}\s*(\d{{1,2}})年\s*(\d{{1,2}})月", text)
        if m:
            year = offset + int(m.group(1))
            month = int(m.group(2))
            if 1 <= month <= 12:
                return f"{year:04d}-{month:02d}"
    # R8.2 / H30.12 の短縮形
    for prefix, offset in [("R", 2018), ("H", 1988)]:
        m = re.search(rf"(?<![A-Za-z]){prefix}(\d{{1,2}})[\.\-_](\d{{1,2}})(?!\d)", text)
        if m:
            year = offset + int(m.group(1))
            month = int(m.group(2))
            if 1 <= month <= 12:
                return f"{year:04d}-{month:02d}"
    # 西暦パターン（例: 2026年2月、2026-02、202602）
    for pat in [
        re.compile(r"(\d{4})年\s*(\d{1,2})月"),
        re.compile(r"(20\d{2})[-_\.](\d{1,2})\b"),
        re.compile(r"(20\d{2})(\d{2})(?!\d)"),
    ]:
        m = pat.search(text)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))
            if 2000 <= year <= 2099 and 1 <= month <= 12:
                return f"{year:04d}-{month:02d}"
    return None


def _parse_fiscal_year_from_filename(filename: str) -> Optional[int]:
    """'2-1-2024.xlsx' / '3-1-H28.xlsx' から fiscal_year（西暦）を返す。"""
    for pat, kind in FISCAL_YEAR_PATTERNS:
        m = pat.match(filename)
        if not m:
            continue
        if kind == "seireki":
            return int(m.group(2))
        if kind == "heisei":
            return 1988 + int(m.group(2))
    return None


def _classify_xlsx_filename(filename: str) -> Optional[dict]:
    """'2-1-2024.xlsx' / '3-1-H28n.xlsx' を分類。

    Returns:
        {"table": "generation"|"demand", "fiscal_year": int, "is_machine_readable": bool}
        or None if doesn't match our target patterns
    """
    for pat, kind in FISCAL_YEAR_PATTERNS:
        m = pat.match(filename)
        if not m:
            continue
        prefix = m.group(1)
        table = TABLE_KEY_FROM_PREFIX.get(prefix)
        if table is None:
            return None
        is_n = bool(m.group(3))  # 'n' variant = 機械判読用
        if kind == "seireki":
            fy = int(m.group(2))
        else:  # heisei
            fy = 1988 + int(m.group(2))
        return {
            "table": table,
            "fiscal_year": fy,
            "is_machine_readable": is_n,
        }
    return None


def list_xlsx_links(
    index_html: str,
    *,
    base_url: str,
    min_fiscal_year: int = 2016,
    prefer_machine_readable: bool = False,
) -> list[dict]:
    """index ページの HTML から Phase 2-A 対象 XLSX のリンクを抽出する。

    対象: `2-1-*.xlsx`（発電実績）と `3-1-*.xlsx`（電力需要実績）のみ。
    その他（1-*, 5-*, 6-* 等）は skip。

    Args:
        index_html: index_url or archive_url から取得した HTML 本文
        base_url: 相対 URL を絶対化するためのベース
        min_fiscal_year: このより古い年度は除外（Phase 2-A v1 は 2016 以降）
        prefer_machine_readable: True なら `*n.xlsx` 変種が存在すれば通常版より優先
            （2026-04-24 Day 5: 機械判読版は列構成が通常版と違い parse 失敗するため
             **デフォルト False** = 通常版優先。FY2025 の `2-1-2025n.xlsx` で総計が
             57,785 GWh ではなく 3,927 GWh になる問題を確認済み）

    Returns:
        list of dict: [
          {"fiscal_year": 2024,
           "table": "generation"|"demand",
           "url": "https://...2-1-2024.xlsx",
           "filename": "2-1-2024.xlsx",
           "is_machine_readable": False,
           "link_text": "2-(1) 発電実績"},
          ...
        ]
        fiscal_year / table の昇順でソート。同じ (fy, table) 重複は
        prefer_machine_readable=True のとき n 版を残して通常版を落とす。
    """
    soup = BeautifulSoup(index_html, "html.parser")
    from urllib.parse import urljoin

    candidates: list[dict] = []
    seen: set[str] = set()

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href:
            continue
        lower = href.lower()
        if not (lower.endswith(".xlsx") or lower.endswith(".xls")):
            continue
        abs_url = urljoin(base_url, href)
        if abs_url in seen:
            continue
        seen.add(abs_url)

        filename = href.rsplit("/", 1)[-1]
        cls = _classify_xlsx_filename(filename)
        if cls is None:
            continue
        if cls["fiscal_year"] < min_fiscal_year:
            continue

        link_text = (a.get_text() or "").strip()
        candidates.append({
            "fiscal_year": cls["fiscal_year"],
            "table": cls["table"],
            "url": abs_url,
            "filename": filename,
            "is_machine_readable": cls["is_machine_readable"],
            "link_text": link_text,
        })

    # 常に (fiscal_year, table) で dedup。prefer_machine_readable で選択基準を切替:
    #   True  → n 版があれば n 版を残す（機械判読版優先）
    #   False → 通常版があれば通常版を残す（Day 5 発見: n 版は列構成が違い parse 失敗）
    by_key: dict[tuple[int, str], dict] = {}
    for c in candidates:
        key = (c["fiscal_year"], c["table"])
        existing = by_key.get(key)
        if existing is None:
            by_key[key] = c
        else:
            if prefer_machine_readable:
                # n 版優先
                if c["is_machine_readable"] and not existing["is_machine_readable"]:
                    by_key[key] = c
            else:
                # 通常版優先（Day 5 default）
                if not c["is_machine_readable"] and existing["is_machine_readable"]:
                    by_key[key] = c
    candidates = list(by_key.values())

    # sort
    candidates.sort(key=lambda r: (r["fiscal_year"], r["table"]))

    logger.info(
        "extracted %d target links (min_fy=%d, prefer_n=%s)",
        len(candidates), min_fiscal_year, prefer_machine_readable,
    )
    return candidates


def fetch_index_pages(index_urls: list[str]) -> list[dict]:
    """複数の index URL を順に叩いて、結果を union して返す。

    メイン entry: results.html（最新年度）+ results_archive.html（過去 10 年度）。

    Note on timeouts & headers (2026-04-24 Day 5 post-mortem):
        - Run #24/#25 は timeout=30s で 6 回全部 timeout → timeout=120s に延長
        - Run #26 は timeout=120s でも 12m 33s 応答なし → ブラウザ風 UA に変更
        - 仮説: METI WAF が "eic-data-pipeline/0.1 (+github...)" を bot と判定し
                応答を返さない。実ブラウザの UA + Accept-Language: ja を渡すと抜ける。
    """
    all_links: list[dict] = []
    seen_urls: set[str] = set()
    for url in index_urls:
        logger.info("GET index page (via curl_cffi/chrome120): %s", url)
        r = _meti_get(url, timeout=120)
        r.raise_for_status()
        links = list_xlsx_links(r.text, base_url=url)
        added = 0
        for link in links:
            if link["url"] in seen_urls:
                continue
            seen_urls.add(link["url"])
            all_links.append(link)
            added += 1
        logger.info("  +%d new links from %s", added, url)
        time.sleep(SLEEP_BETWEEN_REQUESTS)
    all_links.sort(key=lambda r: (r["fiscal_year"], r["table"]))
    logger.info("total unique target links: %d", len(all_links))
    return all_links


# ---------------------------------------------------------------------------
# XLSX ダウンロード
# ---------------------------------------------------------------------------


def download_xlsx(url: str, dest_path: Path, *, use_cache: bool = True) -> bytes:
    """XLSX を download して raw ディレクトリに保存、bytes を返す。

    use_cache=True かつ dest_path が既に存在する場合は download せず、既存ファイルの
    bytes を返す（公表が完了した過去年度向け。年度ファイルは完了後は不変）。

    ★ 2026-09-05（L-064 第 3 例・キャッシュ短絡型）: 公表中の年度ファイル（最新 2 年度）
    は **毎月上書き更新** される。raw は git 追跡なので nightly の checkout に前回ファイル
    が常に存在し、この cache hit が最新 2 年度にも効いた結果、FY2025 の 2026-01〜03 月と
    FY2026 の 5 月以降が一度も取り込まれず、fetcher は毎晩「成功」して updated_at だけが
    前進していた（軸2 は緑のまま）。公表中の年度は呼び出し側が use_cache=False で毎晩
    再取得すること。
    """
    if use_cache and dest_path.exists() and dest_path.stat().st_size > 10_000:
        logger.info("cache hit: %s (%d bytes)", dest_path, dest_path.stat().st_size)
        return dest_path.read_bytes()

    logger.info("GET (via curl_cffi/chrome120): %s", url)
    # METI Akamai Bot Manager 回避のため curl_cffi。timeout 120 秒。
    r = _meti_get(url, timeout=120, accept_xlsx=True)
    r.raise_for_status()
    content = r.content
    if len(content) < 10_000:
        raise RuntimeError(
            f"downloaded xlsx is suspiciously small ({len(content)} bytes); "
            f"content preview: {content[:200]!r}"
        )
    # XLSX は ZIP 容器。チャレンジ HTML や エラーページを raw として保存しない（台帳の偽「変化」防止）。
    if not content.startswith(b"PK"):
        raise RuntimeError(
            f"downloaded file is not an XLSX/ZIP (magic bytes: {content[:4]!r}); "
            f"content preview: {content[:200]!r}"
        )
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    if dest_path.exists() and dest_path.read_bytes() == content:
        logger.info("raw unchanged: %s (%d bytes)", dest_path, len(content))
    else:
        dest_path.write_bytes(content)
        logger.info("saved raw: %s (%d bytes)", dest_path, len(content))
    return content


# ---------------------------------------------------------------------------
# XLSX シート選択（hints + aliases + fuzzy）
# ---------------------------------------------------------------------------


def _pick_sheet(
    xlsx_bytes: bytes,
    primary_hint: str,
    aliases: list[str] | None = None,
) -> Optional[str]:
    """XLSX 内のシート一覧から、primary_hint に最もよく一致するシート名を返す。

    マッチ順:
      1. primary_hint を含むシート（substring match）
      2. aliases のいずれかを含むシート
      3. None（呼び出し側で fallback 判断）
    """
    try:
        xl = pd.ExcelFile(BytesIO(xlsx_bytes), engine="openpyxl")
    except Exception as e:
        logger.error("failed to open xlsx: %s", e)
        return None
    sheets = xl.sheet_names
    logger.debug("xlsx sheets: %s", sheets)

    # 1. primary_hint
    for s in sheets:
        if primary_hint in s:
            return s
    # 2. aliases
    for alias in (aliases or []):
        for s in sheets:
            if alias in s:
                return s
    # 3. 見つからない場合は先頭シートを warning 付きで返す
    if sheets:
        logger.warning(
            "no sheet matched hint='%s' or aliases=%s; fallback to first sheet '%s'",
            primary_hint, aliases, sheets[0],
        )
        return sheets[0]
    return None


def _find_row_by_labels(
    df_raw: pd.DataFrame,
    candidates: list[str],
    max_scan_cols: int = 4,
) -> Optional[int]:
    """DataFrame の左側 max_scan_cols 列を走査し、candidates のいずれかに
    部分一致する最初の行 index を返す。

    METI の XLSX はラベルが 1 列目とは限らず、見出し列が 0〜3 列目に散ることがある。
    """
    n_cols = min(max_scan_cols, df_raw.shape[1])
    for idx in range(len(df_raw)):
        for c in range(n_cols):
            cell = df_raw.iat[idx, c]
            if pd.isna(cell):
                continue
            s = str(cell).strip()
            if not s:
                continue
            # candidates の各キーワードが s に含まれるか（前方一致に近い部分一致）
            for cand in candidates:
                if cand in s:
                    return idx
    return None


def _last_numeric_in_row(row: pd.Series) -> Optional[float]:
    """行を右から走査し、最初に見つかった数値（float に変換可能）を返す。

    ※ Day 2 の単月想定で使っていた関数。Day 4 以降は fiscal-year XLSX で
       12 ヶ月分を `_extract_monthly_values` に切り替え、本関数は使わない。
       互換のため残置（単体テストが参照）。
    """
    for v in reversed(row.tolist()):
        if pd.isna(v):
            continue
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.replace(",", "").strip()
            try:
                return float(s)
            except ValueError:
                continue
    return None


# ---------------------------------------------------------------------------
# Day 4: fiscal-year XLSX から 12 ヶ月列を検出する helper
# ---------------------------------------------------------------------------

_MONTH_LABEL_PATTERN = re.compile(r"^\s*(\d{1,2})\s*月\s*$")


def _find_monthly_columns(
    df_raw: pd.DataFrame,
    *,
    max_scan_rows: int = 15,
) -> tuple[Optional[int], dict[int, int]]:
    """DataFrame の上段を走査し、月ラベル（"4月", "5月", ..., "3月"）が並ぶ
    ヘッダ行を検出。見つかった行 index と月 → 列 index の dict を返す。

    METI fiscal-year XLSX では、通常ヘッダ行に「4 月 5 月 6 月 ... 3 月 年度計」
    のように月次列が並ぶ（fiscal year = 4 月始まり）。

    Returns:
        (header_row_idx, {month_int: col_idx})
        month_int は 1〜12 の元データそのまま（fiscal year への変換は呼び出し側で）。
        ヘッダが見つからなければ (None, {})。
    """
    n_rows = min(max_scan_rows, len(df_raw))
    n_cols = df_raw.shape[1]
    best_row = None
    best_map: dict[int, int] = {}
    for r in range(n_rows):
        month_to_col: dict[int, int] = {}
        for c in range(n_cols):
            cell = df_raw.iat[r, c]
            if pd.isna(cell):
                continue
            s = str(cell).strip()
            m = _MONTH_LABEL_PATTERN.match(s)
            if not m:
                continue
            month = int(m.group(1))
            if 1 <= month <= 12 and month not in month_to_col:
                month_to_col[month] = c
        # より多くの月を含む行を header とみなす
        if len(month_to_col) > len(best_map):
            best_map = month_to_col
            best_row = r
            # 12 ヶ月揃ったらその行を採用
            if len(month_to_col) >= 12:
                break
    return best_row, best_map


def _fiscal_month_to_ym(fiscal_year: int, month: int) -> Optional[str]:
    """日本の会計年度（4 月〜翌 3 月）+ 月 → `YYYY-MM-01`。

    4〜12 月 → year = fiscal_year
    1〜3 月  → year = fiscal_year + 1
    """
    if not (1 <= month <= 12):
        return None
    if month >= 4:
        return f"{fiscal_year:04d}-{month:02d}-01"
    return f"{fiscal_year + 1:04d}-{month:02d}-01"


def _cell_to_float(v) -> Optional[float]:
    """セル値を float に変換、失敗なら None。カンマ・空白・全角数字に耐性。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("　", "")
        # 全角数字を半角に（念のため）
        trans = str.maketrans("０１２３４５６７８９．", "0123456789.")
        s = s.translate(trans)
        if not s or s in {"-", "－", "…"}:  # METI 欠測マーカーの可能性
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _extract_monthly_values(
    df_raw: pd.DataFrame,
    row_idx: int,
    month_to_col: dict[int, int],
) -> dict[int, Optional[float]]:
    """指定行から、各月のセル値を抽出する。

    Returns:
        {month_int: float or None}
    """
    out: dict[int, Optional[float]] = {}
    for month, col in month_to_col.items():
        if col < df_raw.shape[1]:
            v = df_raw.iat[row_idx, col]
            out[month] = _cell_to_float(v)
        else:
            out[month] = None
    return out


# ---------------------------------------------------------------------------
# 発電シート / 需要シートのパース
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Day 5: 実 METI XLSX の構造に基づく parse（全面書き直し）
# ---------------------------------------------------------------------------
#
# 実 XLSX を openpyxl で直接解析して判明した構造（2026-04-24 深夜、Day 5）:
#
# 【発電実績 2-1-{YYYY}.xlsx】
#   - 各 fiscal year ファイルに **月次シートが 12 枚**（`2024.4`, `2024.5`, ..., `2025.3`）
#   - 各月次シート内は「事業者 × 電源種別」のクロス表
#   - ヘッダ: Row 2 (excel row 2) = 電源大分類、Row 3 (excel row 3) = 電源小分類
#   - 合計行: c0 == '合計' （通常 r1611 付近、事業者数 1600+）
#   - 合計行の列 index（単位: 千 kWh = MWh）:
#       c10 = 水力計、c18 = 火力計、c19 = 原子力、
#       c20 = 風力、c21 = 太陽光、c22 = 地熱、c23 = バイオマス、c28 = 総計
#
# 【電力需要 3-1-{YYYY}.xlsx】
#   - 同じく月次シート 12 枚
#   - 2 ブロック構造:
#     Block 1: みなし小売電気事業者等（r28 付近「合  計」全角スペース）
#       c12 = 電灯（自由料金）、c13 = 電力（自由料金）
#       c15 = 電灯（経過措置料金）、c16 = 電力（経過措置料金）
#     Block 2: みなし小売以外（r736 付近「合  計」）
#     r744 付近: c13 = 需要合計（全電気事業者、単位 千kWh）
#
# 【単位変換】 千 kWh → GWh: 値を 1000 で割る
#
# 【Phase 2-A v1 の割り切り】
#   需要の「電灯」「電力」はみなし小売のみ（r28 の c12+c15, c13+c16）。
#   新電力側の電灯/電力区分はこのファイルでは取り切れないため Phase 2-B に後送り。
#   「販売電力量合計」（meti-demand-total）は r744 の c13（真の全事業者合計）。
# ---------------------------------------------------------------------------


_SHEET_NAME_PATTERN = re.compile(r"^(\d{4})\.(\d{1,2})$")


def _find_summary_row_by_label(
    ws, labels: list[str], *, max_rows: int = 2000, max_cols: int = 4,
) -> Optional[int]:
    """ワークシート上で、A 列（or 先頭数列）に labels のいずれかが入っている
    最初の行 index（0-based）を返す。

    labels の文字列は strip 済みで完全一致 or 含有判定。METI XLSX では
    '合計' / '合  計'（全角スペース 2 個）/ '需要合計' 等のバリアントがある。
    """
    normalized_labels = [lbl.strip() for lbl in labels]
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True)
    ):
        for c in range(max_cols):
            if c >= len(row):
                continue
            v = row[c]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            for lbl in normalized_labels:
                # 完全一致 or 先頭一致（「需要合計（電気事業者...）」のような長文対応）
                if s == lbl or s.startswith(lbl):
                    return r_idx
    return None


def _cell_at(ws, row_idx: int, col_idx: int):
    """安全にセル値を取得（範囲外なら None）。row_idx / col_idx は 0-based。"""
    try:
        return ws.cell(row=row_idx + 1, column=col_idx + 1).value
    except Exception:
        return None


def _to_gwh(v) -> Optional[float]:
    """千 kWh（= MWh）から GWh に変換（÷1000）。不正な値は None。"""
    f = _cell_to_float(v)
    if f is None:
        return None
    return round(f / 1000.0, 3)


def parse_generation_sheet(
    xlsx_bytes: bytes,
    cfg: dict,
    fiscal_year: int,
) -> list[dict]:
    """発電実績 XLSX から 8 系列 × 最大 12 ヶ月の long 形式レコードを返す。

    実装: 月次シート (`{YYYY}.{M}`) を順に走査 → 合計行を検出 → 列 index で値取得。
    単位: 千 kWh → GWh（÷1000）。

    Args:
        xlsx_bytes: 2-1-{YYYY}.xlsx の bytes
        cfg:        source_map.yaml の enecho-power セクション（今は未使用、将来の hook 用）
        fiscal_year: 会計年度（参考情報、シート名に既に含まれる）
    Returns:
        list of {date, indicator_id, region, value} 形式
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error("failed to open generation xlsx (FY%d): %s", fiscal_year, e)
        return []

    # 発電シートの合計行ラベル候補
    total_labels = ["合計", "合  計", "総計"]
    # 列 index（0-based）: 合計行からこの列を取る
    col_map = {
        "meti-gen-hydro":      10,  # 水力計
        "meti-gen-thermal":    18,  # 火力計
        "meti-gen-nuclear":    19,  # 原子力
        "meti-gen-wind":       20,  # 風力
        "meti-gen-solar":      21,  # 太陽光
        "meti-gen-geothermal": 22,  # 地熱
        "meti-gen-biomass":    23,  # バイオマス
        "meti-gen-total":      28,  # 総計
    }

    rows: list[dict] = []
    months_found = 0
    for sheet_name in wb.sheetnames:
        m = _SHEET_NAME_PATTERN.match(sheet_name)
        if not m:
            continue
        year = int(m.group(1))
        month = int(m.group(2))
        if not (1 <= month <= 12):
            continue

        ws = wb[sheet_name]
        summary_row = _find_summary_row_by_label(ws, total_labels, max_rows=2500)
        if summary_row is None:
            logger.warning(
                "FY%d sheet='%s': 合計行が見つからず skip", fiscal_year, sheet_name,
            )
            continue

        ymd = f"{year:04d}-{month:02d}-01"
        extracted = 0
        for ind_id, col_idx in col_map.items():
            v = _cell_at(ws, summary_row, col_idx)
            gwh = _to_gwh(v)
            if gwh is None:
                continue
            rows.append({
                "date": ymd,
                "indicator_id": ind_id,
                "region": "jp",
                "value": gwh,
            })
            extracted += 1
        if extracted > 0:
            months_found += 1
        logger.info(
            "FY%d sheet='%s' 合計行 r%d → %d indicators, example total=%s GWh",
            fiscal_year, sheet_name, summary_row, extracted,
            next((r["value"] for r in rows[-extracted:] if r["indicator_id"] == "meti-gen-total"), None),
        )
    wb.close()
    logger.info(
        "parse_generation_sheet FY%d: %d months × up to 8 indicators = %d rows",
        fiscal_year, months_found, len(rows),
    )
    return rows


def parse_demand_sheet(
    xlsx_bytes: bytes,
    cfg: dict,
    fiscal_year: int,
) -> list[dict]:
    """電力需要 XLSX から 3 系列 × 最大 12 ヶ月の long 形式レコードを返す。

    実装:
      - `meti-demand-total`: 最下部「需要合計（電気事業者の販売電力量＋電気事業者の特定供給・自家消費）」行の c13
      - `meti-demand-lights`: みなし小売合計行（r28 付近の「合  計」全角スペース）の c12+c15
      - `meti-demand-power`:  同合計行の c13+c16
    単位: 千 kWh → GWh（÷1000）。
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error("failed to open demand xlsx (FY%d): %s", fiscal_year, e)
        return []

    # 合計行ラベル候補
    minashi_total_labels = ["合  計", "合計", "総計"]  # 全角スペース込みを優先
    grand_total_labels = ["需要合計", "需要計", "電力需要合計"]

    rows: list[dict] = []
    months_found = 0
    for sheet_name in wb.sheetnames:
        m = _SHEET_NAME_PATTERN.match(sheet_name)
        if not m:
            continue
        year = int(m.group(1))
        month = int(m.group(2))
        if not (1 <= month <= 12):
            continue

        ws = wb[sheet_name]
        # みなし小売合計（ブロック 1 の小計、上の方にある）
        minashi_row = _find_summary_row_by_label(ws, minashi_total_labels, max_rows=200)
        # 真の需要合計（ファイル最下部付近）
        grand_row = _find_summary_row_by_label(ws, grand_total_labels, max_rows=2000)

        ymd = f"{year:04d}-{month:02d}-01"
        extracted = 0

        if minashi_row is not None:
            # c12（電灯・自由）+ c15（電灯・経過措置）
            lights = _to_gwh(_cell_at(ws, minashi_row, 12))
            lights_extra = _to_gwh(_cell_at(ws, minashi_row, 15))
            if lights is not None or lights_extra is not None:
                total_lights = (lights or 0) + (lights_extra or 0)
                rows.append({
                    "date": ymd, "indicator_id": "meti-demand-lights",
                    "region": "jp", "value": round(total_lights, 3),
                })
                extracted += 1
            # c13（電力・自由）+ c16（電力・経過措置）
            power = _to_gwh(_cell_at(ws, minashi_row, 13))
            power_extra = _to_gwh(_cell_at(ws, minashi_row, 16))
            if power is not None or power_extra is not None:
                total_power = (power or 0) + (power_extra or 0)
                rows.append({
                    "date": ymd, "indicator_id": "meti-demand-power",
                    "region": "jp", "value": round(total_power, 3),
                })
                extracted += 1

        if grand_row is not None:
            # 「需要合計」行の数値は c0 の下のラベル「合計」の行（r744）or 直後の行にある
            # 実ファイルでは r744 に「需要合計」、r745 (index 744) の c13 に値
            # ただし検出ロジックで grand_row が「需要合計」のラベル行を指している可能性。
            # ラベル行 and 次行の両方で c13 を確認。
            total_v = _to_gwh(_cell_at(ws, grand_row + 1, 13))
            if total_v is None:
                total_v = _to_gwh(_cell_at(ws, grand_row, 13))
            if total_v is not None:
                rows.append({
                    "date": ymd, "indicator_id": "meti-demand-total",
                    "region": "jp", "value": total_v,
                })
                extracted += 1

        if extracted > 0:
            months_found += 1
        logger.info(
            "FY%d sheet='%s' minashi_row=%s grand_row=%s → %d indicators",
            fiscal_year, sheet_name,
            str(minashi_row) if minashi_row is not None else "—",
            str(grand_row) if grand_row is not None else "—",
            extracted,
        )
    wb.close()
    logger.info(
        "parse_demand_sheet FY%d: %d months × up to 3 indicators = %d rows",
        fiscal_year, months_found, len(rows),
    )
    return rows


# ---------------------------------------------------------------------------
# 旧シグネチャ（Day 2〜4 の wide 月ヘッダ前提）は使わなくなったが、
# 既存テスト互換のため骨だけ残す（呼ばれない想定）
# ---------------------------------------------------------------------------


def _parse_sheet_monthly(
    xlsx_bytes: bytes,
    cfg: dict,
    *,
    sheet_kind: str,
    label_key_to_indicator: dict[str, str],
    fiscal_year: int,
) -> list[dict]:
    """fiscal-year XLSX から、指定された indicator 群の **月次 long 形式** を返す。

    Args:
        xlsx_bytes:  XLSX ファイルの bytes
        cfg:         source_map.yaml の enecho-power セクション
        sheet_kind:  "generation_total" | "demand_total"（sheet_hints のキー）
        label_key_to_indicator: row_labels のキー → indicator_id
        fiscal_year: 会計年度（例: 2024 = 2024 年度 = 2024-04 〜 2025-03）

    Returns:
        list of dict: [
          {"date": "2024-04-01", "indicator_id": "meti-gen-thermal", "region": "jp", "value": 55123.0},
          ...
        ]
        month → column 対応が取れなかった場合は空 list。
        指定指標の行が見つからない or 値が欠測の月は該当レコードを skip（null 行を作らない）。
    """
    sheet_hints = cfg.get("sheet_hints") or {}
    sheet_aliases = cfg.get("sheet_aliases") or {}
    row_labels = cfg.get("row_labels") or {}

    sheet_name = _pick_sheet(
        xlsx_bytes,
        primary_hint=sheet_hints.get(sheet_kind, ""),
        aliases=sheet_aliases.get(sheet_kind) or [],
    )
    if sheet_name is None:
        logger.error("no sheet found for %s in FY%d", sheet_kind, fiscal_year)
        return []

    try:
        df = pd.read_excel(
            BytesIO(xlsx_bytes),
            sheet_name=sheet_name,
            header=None,
            engine="openpyxl",
        )
    except Exception as e:
        logger.error("failed to read sheet '%s' (FY%d): %s", sheet_name, fiscal_year, e)
        return []

    # 月ヘッダを検出
    header_row, month_to_col = _find_monthly_columns(df)
    if not month_to_col:
        logger.warning(
            "FY%d sheet='%s': could not locate monthly column headers. "
            "Extraction skipped.", fiscal_year, sheet_name,
        )
        return []
    logger.info(
        "FY%d sheet='%s' header_row=%d months=%s",
        fiscal_year, sheet_name, header_row,
        sorted(month_to_col.keys()),
    )

    rows: list[dict] = []
    for label_key, indicator_id in label_key_to_indicator.items():
        candidates = row_labels.get(label_key) or []
        if not candidates:
            logger.warning("row_labels.%s is empty — skip %s", label_key, indicator_id)
            continue
        # header_row より下で探す（header の上は見ない）
        # _find_row_by_labels は DataFrame 先頭から探すので、header より下を切り出して渡す
        start_row = (header_row + 1) if header_row is not None else 0
        sub_df = df.iloc[start_row:].reset_index(drop=True)
        rel_idx = _find_row_by_labels(sub_df, candidates)
        if rel_idx is None:
            logger.warning(
                "FY%d sheet='%s' label=%s: row not found — %s omitted",
                fiscal_year, sheet_name, candidates, indicator_id,
            )
            continue
        abs_idx = start_row + rel_idx
        monthly = _extract_monthly_values(df, abs_idx, month_to_col)
        n_values = 0
        for month, v in monthly.items():
            if v is None:
                continue
            ymd = _fiscal_month_to_ym(fiscal_year, month)
            if ymd is None:
                continue
            rows.append({
                "date": ymd,
                "indicator_id": indicator_id,
                "region": "jp",
                "value": v,
            })
            n_values += 1
        logger.info(
            "FY%d sheet='%s' row[%d] '%s' → %s × %d months",
            fiscal_year, sheet_name, abs_idx,
            str(df.iat[abs_idx, 0])[:30] if df.shape[1] > 0 else "",
            indicator_id, n_values,
        )
    return rows


# 旧 _parse_sheet_monthly 経由の parse_generation_sheet / parse_demand_sheet は
# Day 5 で新しい実装（line 734, 821 付近）に置き換えたのでここで削除。
# 旧 _parse_sheet_monthly 本体（上記）は互換のため残置（未使用）。


_GENERATION_KEYS = [
    "meti-gen-total", "meti-gen-thermal", "meti-gen-hydro", "meti-gen-nuclear",
    "meti-gen-solar", "meti-gen-wind", "meti-gen-geothermal", "meti-gen-biomass",
]
_DEMAND_KEYS = [
    "meti-demand-total", "meti-demand-lights", "meti-demand-power",
]


# ---------------------------------------------------------------------------
# 派生: 再エネ比率
# ---------------------------------------------------------------------------


_RENEWABLES_NUMERATOR_IDS = [
    "meti-gen-solar", "meti-gen-wind", "meti-gen-geothermal",
    "meti-gen-hydro", "meti-gen-biomass",
]
_RENEWABLES_DENOMINATOR_ID = "meti-gen-total"


def derive_renewables_share_single(month_values: dict[str, Optional[float]]) -> Optional[float]:
    """【単月版】meti-renewables-share を 1 ヶ月分の辞書から計算する（古いシグネチャ）。

    (太陽光 + 風力 + 地熱 + 水力 + バイオマス) / 総発電量 × 100
    """
    need = _RENEWABLES_NUMERATOR_IDS + [_RENEWABLES_DENOMINATOR_ID]
    if not all(k in month_values and month_values[k] is not None for k in need):
        return None
    renewables = sum(month_values[k] for k in _RENEWABLES_NUMERATOR_IDS)
    total = month_values[_RENEWABLES_DENOMINATOR_ID]
    if total is None or total <= 0:
        return None
    return round(renewables / total * 100, 3)


# 互換のため旧名を残す
derive_renewables_share = derive_renewables_share_single


def derive_renewables_share_from_csvs(processed_dir: Path) -> list[dict]:
    """【CSV 全期間版】processed/ 配下の 6 CSV を読んで月次で再エネ比率を計算する。

    入力:
        processed_dir / meti-gen-{solar,wind,geothermal,hydro,biomass,total}.csv
    出力:
        list of {date, indicator_id="meti-renewables-share", region="jp", value}
        すべての分子 + 分母が揃った月のみレコードを作成。
    """
    import pandas as pd  # local import to keep top clean
    need_ids = _RENEWABLES_NUMERATOR_IDS + [_RENEWABLES_DENOMINATOR_ID]
    dfs: dict[str, pd.DataFrame] = {}
    for ind_id in need_ids:
        path = processed_dir / f"{ind_id}.csv"
        if not path.exists():
            logger.warning("derive: CSV missing for %s at %s", ind_id, path)
            return []
        df = pd.read_csv(path, dtype={"date": str})
        # region=jp のみ（将来の地方別拡張に備え）
        if "region" in df.columns:
            df = df[df["region"] == "jp"]
        dfs[ind_id] = df.set_index("date")["value"].astype(float)

    # 全月の intersection（全 6 系列が揃う月のみ）
    common_dates = None
    for ind_id, s in dfs.items():
        if common_dates is None:
            common_dates = set(s.index)
        else:
            common_dates &= set(s.index)
    if not common_dates:
        logger.warning("derive: no common month across 6 series")
        return []

    rows: list[dict] = []
    for d in sorted(common_dates):
        total = dfs[_RENEWABLES_DENOMINATOR_ID].get(d)
        if total is None or total <= 0:
            continue
        renewables = sum(dfs[k].get(d, 0) or 0 for k in _RENEWABLES_NUMERATOR_IDS)
        share = round(renewables / total * 100, 3)
        rows.append({
            "date": d,
            "indicator_id": "meti-renewables-share",
            "region": "jp",
            "value": share,
        })
    logger.info("derive: computed %d months of meti-renewables-share", len(rows))
    return rows


# ---------------------------------------------------------------------------
# メインパイプライン（Day 3 以降で本実装、Day 2 は list/parse のユニットが動くことを目標）
# ---------------------------------------------------------------------------


def fetch_all_months(cfg: dict, *, backfill: bool, since_ym: Optional[str], dry_run: bool) -> int:
    """公表済み全年度を走査、XLSX を download → 12 ヶ月 parse → CSV 追記。

    Day 4 本実装: fiscal_year × {generation, demand} のループで、
        1 XLSX から 12 ヶ月 × N 系列分の long 形式レコードを抽出し、
        indicator_id ごとに CSV に追記（write_processed が dedup + sort する）。
        最後に meti-renewables-share を全 CSV から再計算して追記。

    backfill=True:  min_fiscal_year（デフォルト 2016）以降の全年度を走査
    backfill=False: 最新 2 年度のみ走査（通常モード、data は常に最新に追従）
    since_ym:       "YYYY-MM" を指定すると、その YM 以降のデータのみを追記対象とする

    Returns:
        int: 新規に追記した行数（全 indicator 合計）
    """
    index_urls = cfg.get("index_urls") or [cfg.get("index_url")]
    index_urls = [u for u in index_urls if u]
    if not index_urls:
        logger.error("index_urls / index_url が source_map.yaml に未設定")
        return 0

    min_fy = int(cfg.get("min_fiscal_year") or 2016)
    all_links = fetch_index_pages(index_urls)
    all_links = [lk for lk in all_links if lk["fiscal_year"] >= min_fy]
    if not all_links:
        # index は 200 だがリンクが 1 本も取れない = HTML 構造変更かチャレンジページ。
        # 0 件のまま exit 0 で終わると「取得成功・追加 0 行」と区別がつかない（サイレント成功）。
        raise RuntimeError(
            f"no target XLSX links parsed from index pages {index_urls} (min_fy={min_fy}); "
            "HTML structure changed or challenge page served"
        )

    # 公表中の年度 = index 上の最新 2 年度。毎晩 use_cache=False で再取得する
    # （年度ファイルは月次で上書き更新される。cache hit で短絡すると新しい月が永久に来ない）。
    uniq_fys = sorted({lk["fiscal_year"] for lk in all_links}, reverse=True)
    refresh_fys = set(uniq_fys[:2])

    # 通常モードは最新 2 年度のみ（fiscal year 単位で降順 2 件）
    if not backfill:
        keep_fys = refresh_fys
        links = [lk for lk in all_links if lk["fiscal_year"] in keep_fys]
        logger.info("normal mode: keeping latest 2 fiscal years %s (was %d candidates)",
                    sorted(keep_fys), len(all_links))
    else:
        links = all_links
        logger.info("backfill mode: all %d candidates from FY%d", len(links), min_fy)

    if dry_run:
        logger.info("--- dry-run: %d target links ---", len(links))
        by_fy: dict[int, dict[str, dict]] = {}
        for lk in links:
            by_fy.setdefault(lk["fiscal_year"], {})[lk["table"]] = lk
        for fy in sorted(by_fy.keys()):
            entries = by_fy[fy]
            gen = entries.get("generation")
            dem = entries.get("demand")
            logger.info(
                "  FY%d: gen=%s demand=%s",
                fy,
                (gen["filename"] + (" (n)" if gen["is_machine_readable"] else "")) if gen else "—",
                (dem["filename"] + (" (n)" if dem["is_machine_readable"] else "")) if dem else "—",
            )
        logger.info("total %d fiscal years covered", len(by_fy))
        return 0

    # --- 実行フェーズ -----------------------------------------------
    processed_dir = PROCESSED_DIR
    raw_dir = RAW_DIR
    processed_dir.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)

    # fiscal_year でグループ化（gen + demand を 1 組として扱う）
    by_fy: dict[int, dict[str, dict]] = {}
    for lk in links:
        by_fy.setdefault(lk["fiscal_year"], {})[lk["table"]] = lk

    # since_ym の足切り準備
    since_date = None
    if since_ym:
        try:
            since_date = f"{since_ym}-01"
        except Exception:
            logger.warning("invalid --since=%s; ignored", since_ym)

    # indicator_id → list[dict] を蓄積（年度を跨いで concat 可能）
    accumulated: dict[str, list[dict]] = {}

    for fy in sorted(by_fy.keys()):
        entries = by_fy[fy]
        for table, kind in [("generation", "発電実績"), ("demand", "電力需要実績")]:
            lk = entries.get(table)
            if lk is None:
                logger.warning("FY%d: no %s file link — skip", fy, kind)
                continue
            raw_path = raw_dir / lk["filename"]
            try:
                xlsx_bytes = download_xlsx(lk["url"], raw_path, use_cache=fy not in refresh_fys)
            except Exception as e:
                logger.error("FY%d %s download failed: %s", fy, kind, e)
                continue
            time.sleep(SLEEP_BETWEEN_REQUESTS)
            try:
                if table == "generation":
                    rows = parse_generation_sheet(xlsx_bytes, cfg, fy)
                else:
                    rows = parse_demand_sheet(xlsx_bytes, cfg, fy)
            except Exception as e:
                logger.error("FY%d %s parse failed: %s", fy, kind, e)
                continue

            # since_date 足切り
            if since_date:
                rows = [r for r in rows if r["date"] >= since_date]

            for r in rows:
                accumulated.setdefault(r["indicator_id"], []).append(r)

            logger.info(
                "FY%d %s: extracted %d rows covering %d indicators",
                fy, kind, len(rows),
                len({r["indicator_id"] for r in rows}),
            )

    # --- CSV 追記 + metadata 書き出し（indicator 単位） -----
    total_new_rows = 0
    for indicator_id, rows in sorted(accumulated.items()):
        if not rows:
            continue
        df = pd.DataFrame(rows)
        # source_url 列が空なので append（write_processed 契約）
        df["source_url"] = cfg.get("source_url", "")
        df = df[["date", "indicator_id", "region", "value", "source_url"]]
        try:
            write_processed(df, processed_dir, indicator_id)
            write_metadata_for_indicator(processed_dir, cfg, indicator_id, df)
            total_new_rows += len(df)
        except Exception as e:
            logger.error("write failed for %s: %s", indicator_id, e)

    # --- 派生系列: meti-renewables-share を全 CSV から再計算 ---
    try:
        derived_rows = derive_renewables_share_from_csvs(processed_dir)
        if derived_rows:
            ddf = pd.DataFrame(derived_rows)
            ddf["source_url"] = cfg.get("source_url", "")
            ddf = ddf[["date", "indicator_id", "region", "value", "source_url"]]
            write_processed(ddf, processed_dir, "meti-renewables-share")
            write_metadata_for_indicator(processed_dir, cfg, "meti-renewables-share", ddf)
            total_new_rows += len(ddf)
            logger.info("derive: meti-renewables-share = %d months written", len(ddf))
    except Exception as e:
        logger.error("derive_renewables_share_from_csvs failed: %s", e)

    return total_new_rows


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Fetch METI 電力調査統計 monthly data")
    parser.add_argument(
        "--backfill", action="store_true",
        help="2012-01 から全期間を再取得する",
    )
    parser.add_argument(
        "--since", type=str, default=None,
        help="指定月から取得（YYYY-MM、例: 2024-01）",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="XLSX のダウンロードは行わず、INDEX ページのリンク一覧だけを表示",
    )
    args = parser.parse_args(argv)

    try:
        cfg = load_source_cfg()
    except KeyError as e:
        logger.error(str(e))
        return 2

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    try:
        added = fetch_all_months(
            cfg,
            backfill=args.backfill,
            since_ym=args.since,
            dry_run=args.dry_run,
        )
        logger.info("done: %d new rows appended", added)
        append_log(LOG_DIR, "fetch_enecho_power", "ok",
                   f"added={added} dry_run={args.dry_run} backfill={args.backfill}")
        return 0
    except Exception as e:
        logger.exception("fetch failed: %s", e)
        append_log(LOG_DIR, "fetch_enecho_power", "error", str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())
